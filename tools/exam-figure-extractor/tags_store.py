# -*- coding: utf-8 -*-
"""태그 원천(tags.xlsx / tags.csv) 읽기·쓰기 헬퍼.
build_manifest.py의 열 규약과 동일: A=id, B=tags(쉼표), C=part(선택).
manifest.json은 직접 건드리지 않는다 — build_manifest.py가 이 원천에서 생성한다."""
import csv
import re
import shutil
from datetime import datetime
from pathlib import Path

TAG_SPLIT = re.compile(r"[,;/·\s]+")


def _split(cell):
    if cell is None:
        return []
    return [t for t in TAG_SPLIT.split(str(cell).strip()) if t]


def load_tags(path: Path):
    """반환: {id: {'tags': [...], 'part': [...]}}  (없으면 빈 dict)"""
    path = Path(path)
    if not path.is_file():
        return {}
    out = {}
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            if not key or key.lower() == "id":
                continue
            out[key] = {"tags": _split(row[1] if len(row) > 1 else None),
                        "part": _split(row[2] if len(row) > 2 else None)}
        wb.close()
    else:
        for enc in ("utf-8-sig", "cp949"):
            try:
                with open(path, newline="", encoding=enc) as f:
                    rows = list(csv.reader(f))
                break
            except UnicodeDecodeError:
                continue
        else:
            return {}
        for row in rows:
            if not row or not row[0].strip() or row[0].strip().lower() == "id":
                continue
            out[row[0].strip()] = {
                "tags": _split(row[1] if len(row) > 1 else None),
                "part": _split(row[2] if len(row) > 2 else None)}
    return out


def save_tags(path: Path, data: dict, backup=True):
    """data: {id: {'tags':[...], 'part':[...]}}. 저장 전 자동 백업."""
    path = Path(path)
    if backup and path.is_file():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(path, path.with_suffix(path.suffix + f".{stamp}.bak"))
    ids = sorted(data.keys())
    if path.suffix.lower() == ".xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "tags", "part"])
        for k in ids:
            ws.append([k, ",".join(data[k].get("tags", [])),
                       ",".join(data[k].get("part", []))])
        wb.save(path)
    else:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["id", "tags", "part"])
            for k in ids:
                w.writerow([k, ",".join(data[k].get("tags", [])),
                            ",".join(data[k].get("part", []))])


def default_tags_path(lib_dir: Path):
    """라이브러리의 태그 원천 경로(xlsx 우선, 없으면 csv)."""
    for name in ("tags.xlsx", "tags.csv"):
        p = Path(lib_dir) / name
        if p.is_file():
            return p
    return Path(lib_dir) / "tags.csv"
