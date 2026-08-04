# -*- coding: utf-8 -*-
"""그림 골라내기 — 폴더를 훑으며 O/X 를 찍고 엑셀로 정리한다.

수집한 SVG 수천 장 중 시험지에 쓸 것만 남기려고 만들었다. 한 장씩 보여 주고
O(넣는다) / X(뺀다) 를 누르면 바로 다음 장으로 넘어간다.

## 왜 브라우저인가

SVG 를 빠르게 보려면 브라우저가 사실상 유일하다. tkinter·PyQt 는 SVG 를 래스터로
바꿔야 해서 느리고 의존성이 는다. 대신 엑셀 쓰기는 브라우저가 못 하므로,
**파이썬이 파일을 쓰고 브라우저는 보여 주기만** 한다. tools/png-receiver.py 와 같은 구조다.

## 쓰는 법

    python tools/triage/triage.py                    # 기본 폴더(부품 라이브러리)
    python tools/triage/triage.py <폴더> [포트]

브라우저가 자동으로 열린다. 키보드:
    O 넣는다 · X 뺀다 · S 보류 · Z 되돌리기

## 나오는 것

    _work/triage/triage.xlsx   엑셀 (파일명 · 판정 …)
    _work/triage/marks.json    진행 상태 — 껐다 켜도 이어서 한다

엑셀은 25장마다, 그리고 끝낼 때 자동으로 다시 쓴다.
"""

import io
import json
import mimetypes
import re
import pathlib
import sys
import threading
import time
import webbrowser

sys_path_added = True
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs, unquote

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
import picker

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)

HERE = pathlib.Path(__file__).resolve().parent
ROOT = HERE.parent.parent                       # 5E_main
DEFAULT_DIR = ROOT / "assets" / "parts-library" / "svg"
OUT_DIR = ROOT / "_work" / "triage"

TARGET = pathlib.Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_DIR
PORT = int(sys.argv[2]) if len(sys.argv) > 2 else 8595

MARKS = OUT_DIR / "marks.json"
XLSX = OUT_DIR / "triage.xlsx"
SCORES = OUT_DIR / "scores.json"
ROUNDS = OUT_DIR / "rounds.json"
CLUSTERS = OUT_DIR / "clusters.json"     # 닮은 그림 무리 (지문으로 묶은 것)

# 기준이 되는 레퍼런스 — 손으로 고른 것들. 라운드 화면 맨 위에 늘 띄운다.
REF_FILES = ["b_animal_cell.svg", "b_mitochondrion.svg", "b_prokaryote.svg",
             "b_plant_cell.svg", "b_flower.svg", "c_distillation.svg",
             "p_pendulum_coord.svg", "p_pendulum.svg",
             "p_pendulum_height.svg", "p_ohms_law.svg"]

# ── 상태 ────────────────────────────────────────────────────────────────
lock = threading.Lock()
marks = {}          # 파일명 -> {"mark": "O"|"X"|"S", "at": ISO}
order = []          # 파일명 목록 (정렬 고정)
meta = {}           # 파일명 -> 수집 정보 (harvest.json 이 있으면)
scores = {}         # 파일명 -> 실제로 그려 보고 잰 값 (autoscan.html 이 넣는다)
_dirty = 0


# ── 등급 ────────────────────────────────────────────────────────────────
# 두 가지를 함께 본다. 어느 하나만으로는 안 된다:
#   · 제목만 보면 문장(紋章) 그림이 그냥 통과한다.
#   · 픽셀만 보면 화학 구조식과 세포 그림을 못 가른다.
#
# 기준은 손으로 고른 10장을 실제로 재서 잡았다:
#   ink(잉크가 덮은 넓이) 0.05~0.36 · 도형 12~599개
# 처음엔 "검은 선 비율"을 쓰려 했는데, 128px 로 줄이면 가는 검은 선이 흰색과
# 섞여 사라진다(기준 10장의 dark 가 거의 전부 0 이었다). 그래서 버렸다.
#
# ⚠ 이건 거친 1차 체다. "좋은 그림"을 골라 주는 게 아니라 **확실히 아닌 것을
#    빼 주는** 것이다. A 를 받은 것도 사람이 눈으로 봐야 한다.

HERALDRY = re.compile(r"blason|coat of arms|wappen|escudo|armoiries|герб|stemma|crest of", re.I)
# 화학 구조식은 이미지로 넣을 게 아니라 그리는 도구로 만들 것이다(교사 지시).
# 그래서 후보에서 아예 뺀다. 약품·화합물 이름의 흔한 꼴을 함께 잡는다.
CHEMFORM = re.compile(
    r"\d[,\-]\d|-2D\b|skeletal|structural formula|struttura|säure|"
    r"\bacide\b|zuur|smiles|\bchemical structure\b|"
    r"\bCoA\b|\bsynthes(e|is)\b|biosynthesis|\bpathway\b|"
    r"\b\d+-[A-Z]{2,4}\b|\b\d+[A-Z]-[A-Z0-9]|"        # 2-MGA · 2C-G-4 · 3-F-BPAP 꼴
    r"\b[A-Za-z]{5,}(ine|amide|oate|ylate|osine|azine|azole|idine)\b|"
    # "…ic acid" 만 잡는다. 맨 acid 를 잡으면 "Acid rain"(산성비) 까지 날아간다.
    r"\b\w+ic acid\b|\b\w+säure\b|\b\w+saeure\b",
    re.I)
STANDARD = re.compile(r"ГОСТ|GOST|DIN \d|ISO \d|IEC \d{4}|заводское|клеймо|"
                      r"iso[_ ]?7010|pictogram", re.I)
# 회로도도 뺀다 — 5E 는 회로를 add_circuit 도구로 그린다(교사 지시).
# 15라운드에서 회로도 11장이 모두 X 였다.
CIRCUIT = re.compile(r"circuit|schaltung|amplifier|transistor|\bdiode\b|\bbias\b|"
                     r"oscillator|rectifier|\bop-?amp\b|logic gate|flip-?flop|"
                     r"schematic|wiring|directform|biquad|\bresistor\b|capacitor|"
                     r"\bkondensator\b|\bwiderstand\b|파형|회로", re.I)
LANGSFX = re.compile(r"[-_ ]([a-z]{2}|[a-z]{2}[-_][A-Za-z]{2})$")
DRAW_RE = re.compile(r"<(path|circle|ellipse|rect|polygon|polyline|line)\b")

static = {}     # 파일명 -> {"draw": n, "flag": str|None, "dup": bool}


def build_static():
    """제목·복잡도처럼 그려 보지 않아도 아는 것을 미리 잰다."""
    seen = {}
    for name in order:
        it = meta.get(name, {})
        title = it.get("sourceTitle") or it.get("name", "") or name
        f = TARGET / name
        try:
            n = len(DRAW_RE.findall(f.read_text(encoding="utf-8", errors="replace")))
        except Exception:
            n = 0
        flag = None
        if HERALDRY.search(title):
            flag = "문장·기장 그림"
        elif STANDARD.search(title):
            flag = "규격·표지 기호"
        elif CHEMFORM.search(title):
            flag = "화학 구조식(도구로 그림)"
        elif it.get("part") == "회로" or CIRCUIT.search(title):
            flag = "회로도(도구로 그림)"
        # 같은 그림의 다른 언어판 — 첫 장만 남기고 나머지는 중복으로 본다
        base = LANGSFX.sub("", title).strip().lower()
        dup = base in seen
        seen.setdefault(base, name)
        static[name] = {"draw": n, "flag": flag, "dup": dup}


def grade_of(name_or_score, name=None):
    s = name_or_score if isinstance(name_or_score, dict) or name_or_score is None else None
    st = static.get(name or "", {})
    if s is None and name:
        s = scores.get(name)
    if s is None:
        return "?", "안 잼"
    if s.get("fail"):
        return "F", "안 그려짐"

    ink = s.get("ink", 0)
    draw = st.get("draw", 0)

    if ink < 0.015:
        return "C", "거의 빔"
    if ink > 0.60:
        return "C", "색면이 꽉 참"
    if draw and draw < 8:
        return "C", "너무 단순(기호)"
    if st.get("flag"):
        return "C", st["flag"]
    if st.get("dup"):
        return "B", "다른 언어판 중복"
    if draw > 900:
        return "B", "너무 복잡(지도·차트)"
    return "A", "쓸 만함"


def load():
    global order, marks, meta
    order = sorted(p.name for p in TARGET.glob("*.svg"))
    if not order:
        order = sorted(p.name for p in TARGET.iterdir()
                       if p.suffix.lower() in (".svg", ".png", ".jpg", ".jpeg", ".webp"))
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if MARKS.exists():
        marks = json.loads(MARKS.read_text(encoding="utf-8"))
    if SCORES.exists():
        scores.update(json.loads(SCORES.read_text(encoding="utf-8")))
    _later_build_static = True
    # 수집 정보가 있으면 화면에 같이 띄운다 — 라이선스·출처를 보고 판단할 수 있게
    h = TARGET.parent / "harvest.json"
    if h.exists():
        for it in json.loads(h.read_text(encoding="utf-8")):
            meta[it.get("file", "")] = it
    build_static()
    load_clusters()


def save_marks():
    MARKS.write_text(json.dumps(marks, ensure_ascii=False, indent=1), encoding="utf-8")


def save_xlsx():
    """엑셀로 정리. 첫 두 칸이 파일명과 판정이고, 나머지는 판단에 쓰라고 덧붙인 것."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  ⚠ openpyxl 이 없어 엑셀을 못 쓴다 — pip install openpyxl")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "선별"
    head = ["파일명", "판정", "자동등급", "등급사유", "무리대표", "무리크기",
            "이름", "과목", "분류", "라이선스", "출처", "태그", "판정시각"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F6F8FA")
    for name in order:
        m = marks.get(name, {})
        it = meta.get(name, {})
        g, why = grade_of(scores.get(name), name)
        rep = rep_of.get(name, name)
        ws.append([
            name,
            m.get("mark", ""),
            g,
            why,
            "대표" if rep == name else rep,
            len(clusters.get(rep, [name])),
            it.get("name", ""),
            it.get("subjectLabel", ""),
            it.get("part", ""),
            it.get("license", ""),
            it.get("source", ""),
            ", ".join(it.get("sourceTags", [])[:6]),
            m.get("at", ""),
        ])
    widths = [44, 6, 9, 20, 34, 8, 26, 7, 11, 16, 52, 40, 20]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].alignment = Alignment(horizontal="center")
    wb.save(XLSX)


def load_rounds():
    return json.loads(ROUNDS.read_text(encoding="utf-8")) if ROUNDS.exists() else []


def round_verdicts(rounds):
    """라운드에서 받은 O/X 를 모은다. 레퍼런스 10장은 늘 accept 쪽에 둔다."""
    acc = [f for f in REF_FILES if (TARGET / f).exists()]
    rej = []
    for r in rounds:
        for f, v in r.items():
            (acc if v == "O" else rej).append(f)
    return acc, rej


def counts():
    c = {"O": 0, "X": 0, "S": 0}
    for m in marks.values():
        c[m.get("mark", "S")] = c.get(m.get("mark", "S"), 0) + 1
    return c


def grade_counts():
    c = {}
    for f in order:
        g = grade_of(scores.get(f), f)[0]
        c[g] = c.get(g, 0) + 1
    return c


# ── 닮은 그림 묶음 ──────────────────────────────────────────────────────
# 제목이 전혀 달라도 그림이 같은 것들이 있다("Measuring cylinder" 와 "Mischzylinder").
# 그래서 그림 지문으로 묶고, 무리마다 **대표 한 장만** 보여 준다.
clusters = {}        # 대표 -> [식구들]
rep_of = {}          # 파일 -> 대표


def load_clusters():
    if not CLUSTERS.exists():
        return
    clusters.update(json.loads(CLUSTERS.read_text(encoding="utf-8")))
    for lead, members in clusters.items():
        for f in members:
            rep_of[f] = lead


def build_queue():
    """이제 볼 것만 남긴 줄. A 등급 · 무리 대표 · 아직 안 정한 것."""
    q, seen_rep = [], set()
    for f in order:
        if f in marks:
            seen_rep.add(rep_of.get(f, f))      # 무리 식구 중 하나라도 정했으면 그 무리는 끝
    for f in order:
        if f in marks:
            continue
        if grade_of(scores.get(f), f)[0] != "A":
            continue
        r = rep_of.get(f, f)
        if r in seen_rep:
            continue
        seen_rep.add(r)
        q.append(f)
    return q


# ── 서버 ────────────────────────────────────────────────────────────────
class H(BaseHTTPRequestHandler):
    def log_message(self, *a):
        pass

    def _send(self, code, body, ctype="application/json; charset=utf-8"):
        if isinstance(body, str):
            body = body.encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        u = urlparse(self.path)
        path = unquote(u.path)

        if path in ("/", "/index.html"):
            return self._send(200, (HERE / "index.html").read_bytes(),
                              "text/html; charset=utf-8")

        if path == "/autoscan":
            return self._send(200, (HERE / "autoscan.html").read_bytes(),
                              "text/html; charset=utf-8")

        if path == "/rounds":
            return self._send(200, (HERE / "rounds.html").read_bytes(),
                              "text/html; charset=utf-8")

        if path == "/api/round":
            with lock:
                rounds = load_rounds()
                acc, rej = round_verdicts(rounds)
                cand = picker.pick(order, meta, static, scores,
                                   accepted=acc, rejected=rej, n=10)
                return self._send(200, json.dumps({
                    "round": len(rounds) + 1,
                    "refs": [{"file": r, "title": (meta.get(r) or {}).get("sourceTitle", r)}
                             for r in REF_FILES if (TARGET / r).exists()],
                    "items": [{"file": f, "score": s, "why": w,
                               "title": (meta.get(f) or {}).get("sourceTitle") or f,
                               "part": (meta.get(f) or {}).get("part", ""),
                               "subject": (meta.get(f) or {}).get("subjectLabel", ""),
                               "license": (meta.get(f) or {}).get("license", "")}
                              for f, s, w in cand],
                    "history": [{"round": i + 1,
                                 "O": sum(1 for v in r.values() if v == "O"),
                                 "X": sum(1 for v in r.values() if v == "X")}
                                for i, r in enumerate(rounds)],
                }, ensure_ascii=False))

        if path == "/api/unscored":
            with lock:
                return self._send(200, json.dumps(
                    {"files": [f for f in order if f not in scores],
                     "done": len(scores), "total": len(order)}))

        if path == "/api/list":
            with lock:
                return self._send(200, json.dumps({
                    "dir": str(TARGET),
                    "total": len(order),
                    "files": order,
                    "marks": marks,
                    "meta": {k: {"name": v.get("name", ""),
                                 "subjectLabel": v.get("subjectLabel", ""),
                                 "part": v.get("part", ""),
                                 "license": v.get("license", ""),
                                 "source": v.get("source", ""),
                                 "tags": v.get("sourceTags", [])[:8]}
                             for k, v in meta.items()},
                    "counts": counts(),
                    "queue": build_queue(),
                    "clusterSize": {f: len(clusters.get(rep_of.get(f, f), [f]))
                                    for f in order},
                    "grades": {f: grade_of(scores.get(f), f)[0] for f in order},
                    "reasons": {f: grade_of(scores.get(f), f)[1] for f in order},
                    "gradeCounts": grade_counts(),
                    "xlsx": str(XLSX),
                }, ensure_ascii=False))

        if path.startswith("/img/"):
            name = path[5:]
            f = TARGET / name
            # 폴더 밖으로 새어 나가지 않게
            if not f.resolve().is_relative_to(TARGET) or not f.exists():
                return self._send(404, b"no", "text/plain")
            ctype = mimetypes.guess_type(name)[0] or "application/octet-stream"
            if name.lower().endswith(".svg"):
                ctype = "image/svg+xml"
            return self._send(200, f.read_bytes(), ctype)

        return self._send(404, b"no", "text/plain")

    def do_POST(self):
        global _dirty
        u = urlparse(self.path)
        n = int(self.headers.get("Content-Length") or 0)
        raw = self.rfile.read(n).decode("utf-8") if n else "{}"

        if u.path == "/api/mark":
            d = json.loads(raw)
            name, mark = d.get("file"), d.get("mark")
            with lock:
                if mark in ("O", "X", "S"):
                    marks[name] = {"mark": mark,
                                   "at": time.strftime("%Y-%m-%d %H:%M:%S")}
                elif mark is None:
                    marks.pop(name, None)          # 되돌리기
                save_marks()
                _dirty += 1
                if _dirty >= 25:
                    save_xlsx(); _dirty = 0
                return self._send(200, json.dumps({"ok": True, "counts": counts()},
                                                  ensure_ascii=False))

        if u.path == "/api/round":
            d = json.loads(raw)
            with lock:
                rounds = load_rounds()
                rounds.append(d.get("verdicts", {}))
                ROUNDS.write_text(json.dumps(rounds, ensure_ascii=False, indent=1),
                                  encoding="utf-8")
                acc, rej = round_verdicts(rounds)
                return self._send(200, json.dumps(
                    {"ok": True, "round": len(rounds),
                     "accepted": len(acc), "rejected": len(rej)}))

        if u.path == "/api/score":
            d = json.loads(raw)
            with lock:
                scores.update(d.get("scores", {}))
                SCORES.write_text(json.dumps(scores, ensure_ascii=False), encoding="utf-8")
                return self._send(200, json.dumps(
                    {"ok": True, "done": len(scores), "grades": grade_counts()}))

        if u.path == "/api/save":
            with lock:
                save_marks(); save_xlsx(); _dirty = 0
            return self._send(200, json.dumps({"ok": True, "xlsx": str(XLSX)},
                                              ensure_ascii=False))

        return self._send(404, b"no", "text/plain")


def main():
    if not TARGET.exists():
        print(f"폴더가 없다: {TARGET}"); return 1
    load()
    if not order:
        print(f"그림이 없다: {TARGET}"); return 1
    c = counts()
    print(f"폴더   {TARGET}")
    print(f"그림   {len(order)}장 (이미 판정 {len(marks)}장 — O {c['O']} · X {c['X']} · 보류 {c['S']})")
    print(f"엑셀   {XLSX}")
    print(f"주소   http://127.0.0.1:{PORT}/")
    print("끝낼 때는 이 창에서 Ctrl+C — 그 전에 엑셀은 자동 저장된다\n")
    srv = ThreadingHTTPServer(("127.0.0.1", PORT), H)
    threading.Timer(0.6, lambda: webbrowser.open(f"http://127.0.0.1:{PORT}/")).start()
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        with lock:
            save_marks(); save_xlsx()
        print(f"\n저장 완료 → {XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
