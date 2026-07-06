# ===== EXAM LIBRARY MANIFEST BUILDER =====
#
# assets/exam-library/images/ 의 PNG 파일명을 스캔해 검색용 manifest.json을 생성한다.
# 태그는 assets/exam-library/tags.xlsx (또는 tags.csv)에서 병합하고,
# tag-vocab.json 어휘집에 없는 태그는 경고로 보고한다.
#
# 실행:  python scripts/build_manifest.py
# 흐름:  이미지 추가/삭제 또는 tags.xlsx 수정 → 이 스크립트 재실행 → 앱 새로고침
#
# 파일명 규칙: <과목코드>_<학년도4자리>_<시험월2자리>_<문항번호>.png
#   예) p1_2026_11_01.png  →  2026학년도 수능 물리1 1번
#   프리픽스 뒤에 붙는 라벨은 무시된다:
#   예) "p1_2025_03_02 [2025학년도 3월 학평 물리1 2번].png" 도 동일하게 인식

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
LIB_DIR = ROOT / "assets" / "exam-library"

SUBJECT_MAP = {
    "p1": "물리1", "p2": "물리2",
    "c1": "화학1", "c2": "화학2",
    "b1": "생명1", "b2": "생명2",
    "e1": "지구1", "e2": "지구2",
    "i1": "통합과학",
}

# 파일명 stem 앞부분만 매칭 — 뒤에 붙는 " [라벨]" 등은 무시
FILENAME_RE = re.compile(r"^([a-z]+\d*)_(\d{4})_(\d{2})_(\d{1,3})", re.IGNORECASE)

# 태그 구분자: 쉼표/세미콜론/슬래시/가운뎃점/공백 모두 허용
TAG_SPLIT_RE = re.compile(r"[,;/·\s]+")


def exam_label(month: int) -> str:
    if month == 11:
        return "수능"
    if month in (6, 9):
        return f"{month}월 모평"
    return f"{month}월 학평"


def load_vocab(path: Path):
    """어휘집 로드 → (전체 태그 set, 카테고리 리스트). 파일이 없으면 검증 생략."""
    if not path.is_file():
        return None, []
    data = json.loads(path.read_text(encoding="utf-8"))
    categories = data.get("categories", [])
    all_tags = {t for c in categories for t in c.get("tags", [])}
    return all_tags, [{"name": c["name"], "tags": c["tags"]} for c in categories]


def parse_tag_cell(cell) -> list:
    if cell is None:
        return []
    return [t for t in TAG_SPLIT_RE.split(str(cell).strip()) if t]


def load_tags_xlsx(path: Path) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    tags_by_id = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if not key or key.lower() == "id":  # 헤더 행
            continue
        tags_by_id[key] = parse_tag_cell(row[1] if len(row) > 1 else None)
    wb.close()
    return tags_by_id


def load_tags_csv(path: Path) -> dict:
    import csv
    for enc in ("utf-8-sig", "cp949"):
        try:
            with open(path, newline="", encoding=enc) as f:
                rows = list(csv.reader(f))
            break
        except UnicodeDecodeError:
            continue
    else:
        raise SystemExit(f"오류: {path} 인코딩을 읽을 수 없습니다 (utf-8/cp949 아님)")
    tags_by_id = {}
    for row in rows:
        if not row or not row[0].strip() or row[0].strip().lower() == "id":
            continue
        tags_by_id[row[0].strip()] = parse_tag_cell(",".join(row[1:]))
    return tags_by_id


def main():
    ap = argparse.ArgumentParser(description="기출 이미지 라이브러리 manifest.json 생성")
    ap.add_argument("--images", type=Path, default=LIB_DIR / "images")
    ap.add_argument("--tags", type=Path, default=None,
                    help="태그 시트 경로 (기본: exam-library/tags.xlsx 또는 tags.csv)")
    ap.add_argument("--vocab", type=Path, default=LIB_DIR / "tag-vocab.json")
    ap.add_argument("--out", type=Path, default=LIB_DIR / "manifest.json")
    args = ap.parse_args()

    if not args.images.is_dir():
        raise SystemExit(f"오류: 이미지 폴더가 없습니다 → {args.images}")

    # ----- 태그 시트 로드 -----
    tags_path = args.tags
    if tags_path is None:
        for candidate in (LIB_DIR / "tags.xlsx", LIB_DIR / "tags.csv"):
            if candidate.is_file():
                tags_path = candidate
                break
    tags_by_id = {}
    if tags_path and tags_path.is_file():
        if tags_path.suffix.lower() == ".xlsx":
            tags_by_id = load_tags_xlsx(tags_path)
        else:
            tags_by_id = load_tags_csv(tags_path)

    vocab_set, vocab_categories = load_vocab(args.vocab)

    # ----- 이미지 스캔 -----
    items = []
    skipped = []          # (파일명, 사유)
    seen_ids = {}         # id → 파일명 (중복 감지)
    unknown_tags = {}     # 태그 → [id...]

    for f in sorted(args.images.iterdir()):
        if not f.is_file() or f.suffix.lower() != ".png":
            if f.name != ".gitkeep":
                skipped.append((f.name, "png 아님"))
            continue
        m = FILENAME_RE.match(f.stem)
        if not m:
            skipped.append((f.name, "파일명 규칙 불일치 (과목_학년도_월_번호)"))
            continue
        subject_code = m.group(1).lower()
        year, month, no = int(m.group(2)), int(m.group(3)), int(m.group(4))
        if not (1 <= month <= 12):
            skipped.append((f.name, f"시험월 {month:02d} 비정상"))
            continue
        item_id = f"{subject_code}_{year}_{month:02d}_{no:02d}"
        if item_id in seen_ids:
            skipped.append((f.name, f"id 중복 (동일 문항: {seen_ids[item_id]})"))
            continue
        seen_ids[item_id] = f.name

        subject_label = SUBJECT_MAP.get(subject_code, subject_code)
        exam = exam_label(month)
        tags = tags_by_id.get(item_id, [])
        if vocab_set is not None:
            for t in tags:
                if t not in vocab_set:
                    unknown_tags.setdefault(t, []).append(item_id)

        items.append({
            "id": item_id,
            "file": f.name,
            "subject": subject_code,
            "subjectLabel": subject_label,
            "year": year,
            "month": month,
            "exam": exam,
            "no": no,
            "title": f"{year}학년도 {exam} {subject_label} {no}번",
            "tags": tags,
        })

    items.sort(key=lambda x: (x["subject"], -x["year"], x["month"], x["no"]))

    manifest = {
        "version": "exam-library-v1",
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "count": len(items),
        "tagVocab": vocab_categories,
        "items": items,
    }
    args.out.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=1), encoding="utf-8")

    # ----- 보고 -----
    tagged = sum(1 for it in items if it["tags"])
    print(f"manifest.json 생성 완료 → {args.out}")
    print(f"  문항 {len(items)}개 (태그 있음 {tagged} / 없음 {len(items) - tagged})")
    if tags_path and tags_path.is_file():
        orphan_ids = sorted(set(tags_by_id) - set(seen_ids))
        print(f"  태그 시트: {tags_path.name} ({len(tags_by_id)}행)")
        if orphan_ids:
            print(f"  [경고] 태그 시트에 있으나 이미지가 없는 id {len(orphan_ids)}개:")
            for i in orphan_ids[:10]:
                print(f"    - {i}")
            if len(orphan_ids) > 10:
                print(f"    … 외 {len(orphan_ids) - 10}개")
    else:
        print("  태그 시트 없음 → 태그 없이 생성 (문항번호 검색만 가능)")
    if skipped:
        print(f"  [경고] 건너뛴 파일 {len(skipped)}개:")
        for name, reason in skipped:
            print(f"    - {name}: {reason}")
    if unknown_tags:
        print(f"  [경고] 어휘집(tag-vocab.json)에 없는 태그 {len(unknown_tags)}종:")
        for t, ids in sorted(unknown_tags.items()):
            print(f"    - {t} ({len(ids)}개 문항, 예: {ids[0]})")
        print("    → 오타면 tags.xlsx 수정, 새 개념이면 tag-vocab.json에 추가 후 재실행")


if __name__ == "__main__":
    main()
